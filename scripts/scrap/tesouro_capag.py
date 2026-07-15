#!/usr/bin/env python3
"""
Fetch CAPAG (Capacidade de Pagamento) fiscal-capacity ratings for Brazilian
states and municipalities -> Parquet -> beelink.

Source: Tesouro Transparente's CKAN instance (tesourotransparente.gov.br/ckan),
datasets "capag-estados" and "capag-municipios". Discovered via CKAN
package_show (https://www.tesourotransparente.gov.br/ckan/api/3/action/package_show?id=capag-estados
and ?id=capag-municipios), which lists direct-download resource URLs — no
auth, no scraping of the interactive portal needed.

Two tables, different fetch strategy per table:

  - estados: one flat CSV per year (2018-2025), schema is stable enough to
    concatenate directly with a `ano_referencia`/`arquivo_origem` column.
    Older years (2018-2020) lack the Nota1/2/3 and "Qualidade da informacao
    contabil" columns present from ~2021 onward -- left null for those years
    rather than forcing a fake value.

  - municipios: NOT flat per-year CSVs. Tesouro publishes CAPAG municipios
    as a single large analytical XLSX workbook (one file per publication
    date, most recent covers ano-base 2025) with dozens of raw-accounting
    sheets per fiscal year, multi-row headers, and revision/formula columns
    -- not something to safely diff/reconcile across the ~16 historical
    file revisions in this pass. Instead this pulls just the "Previa da
    CAPAG" sheet (data_only=True to get openpyxl's cached formula results,
    since the live formulas reference other sheets) from the single latest
    workbook -- a clean final-rating snapshot per municipio (5,568 rows,
    matching Brazil's municipio count), same shape as the estados table.
    Historical municipio-level CAPAG could be added later by parsing each
    year's raw sheets individually if needed.

Usage:
    python3 scripts/scrap/tesouro_capag.py
"""

import csv
import io
import json
import subprocess
import sys
import time
from pathlib import Path
from urllib.request import Request, urlopen

BEELINK_HOST = "beelink"
DATASET_PATH = "~/rodado/br_tesouro_capag"
BEELINK_PATH = f"{DATASET_PATH}/estados"
UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"
TEMP_DIR = Path("/private/tmp/claude-501/-Users-polux-Projetos-rodado/c780c9c0-b6b3-44b0-964e-08a3b2f2024c/scratchpad/tesouro_capag")

CKAN_BASE = "https://www.tesourotransparente.gov.br/ckan"
FETCH_TIMEOUT = 60


def http_get(url: str, timeout: int = FETCH_TIMEOUT) -> bytes:
    req = Request(url, headers={"User-Agent": UA})
    with urlopen(req, timeout=timeout) as resp:
        return resp.read()


def ckan_resources(dataset_id: str) -> list:
    url = f"{CKAN_BASE}/api/3/action/package_show?id={dataset_id}"
    data = json.loads(http_get(url))
    return data["result"]["resources"]


def fetch_estados_rows() -> list:
    print("=== estados ===")
    resources = ckan_resources("capag-estados")
    rows = []
    for r in resources:
        if r.get("format", "").upper() != "CSV":
            continue
        name = r["name"]
        url = r["url"]
        print(f"  fetching {name} ...")
        raw = http_get(url)
        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text), delimiter=";")
        for row in reader:
            clean = {k.strip(): (v.strip() if isinstance(v, str) else v) for k, v in row.items() if k}
            clean["arquivo_origem"] = name
            rows.append(clean)
    print(f"  Total estados rows: {len(rows)}")
    return rows


def fetch_municipios_rows() -> list:
    import openpyxl

    print("=== municipios ===")
    resources = ckan_resources("capag-municipios")
    xlsx_resources = [r for r in resources if r.get("format", "").upper() == "XLSX"]
    if not xlsx_resources:
        print("  No XLSX resources found — aborting municipios.", file=sys.stderr)
        return []
    # CKAN lists resources in creation order; the last one is the most recent
    # publication (confirmed: last entry is the highest year/date in the name).
    latest = xlsx_resources[-1]
    name = latest["name"]
    url = latest["url"]
    print(f"  fetching latest workbook: {name} ...")

    xlsx_path = TEMP_DIR / "capag_municipios_latest.xlsx"
    TEMP_DIR.mkdir(parents=True, exist_ok=True)
    raw = http_get(url, timeout=180)
    xlsx_path.write_bytes(raw)
    print(f"  downloaded {xlsx_path.stat().st_size / 1e6:.1f} MB")

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    sheet_name = next((s for s in wb.sheetnames if "Prévia" in s or "Previa" in s), None)
    if sheet_name is None:
        print(f"  Could not find 'Previa da CAPAG' sheet among {wb.sheetnames} — aborting.", file=sys.stderr)
        return []
    ws = wb[sheet_name]

    header = None
    rows = []
    for r in ws.iter_rows(values_only=True):
        if header is None:
            if r and r[0] == "Código Município Completo":
                header = [h for h in r if h is not None]
            continue
        if r[0] is None:
            continue
        row_dict = dict(zip(header, r[: len(header)]))
        # This workbook mixes types within several columns across its 5,568
        # rows (e.g. "Indicador N" is usually a float but holds the sentinel
        # string "n.d." for some municipios; other columns mix int/float/str
        # in ways not worth enumerating one by one). Stringify everything
        # except the numeric município code so pyarrow always gets a single
        # consistent type per column.
        code = row_dict.get("Código Município Completo")
        row_dict = {
            k: (str(v) if v is not None else None)
            for k, v in row_dict.items()
        }
        row_dict["Código Município Completo"] = code
        row_dict["arquivo_origem"] = name
        rows.append(row_dict)

    print(f"  Total municipios rows: {len(rows)}")
    return rows


def write_and_push(rows: list, table_name: str) -> bool:
    import pyarrow as pa
    import pyarrow.parquet as pq

    if not rows:
        print(f"  No rows for {table_name} — skipping push.", file=sys.stderr)
        return False

    table = pa.Table.from_pylist(rows)
    parquet_path = TEMP_DIR / f"{table_name}.parquet"
    pq.write_table(table, str(parquet_path), compression="zstd")
    print(f"  Wrote {parquet_path} ({parquet_path.stat().st_size / 1e6:.2f} MB, {table.num_rows} rows)")

    beelink_path = f"{DATASET_PATH}/{table_name}"
    for attempt in range(20):
        mkdir_ok = subprocess.run(f"ssh {BEELINK_HOST} 'mkdir -p {beelink_path}'", shell=True).returncode == 0
        if mkdir_ok:
            break
        print(f"  ssh unreachable (attempt {attempt + 1}/20), retrying in 15s ...", file=sys.stderr)
        time.sleep(15)
    else:
        print(f"ssh beelink unreachable after retries — aborting push for {table_name}.", file=sys.stderr)
        return False
    result = subprocess.run(
        f"rsync -av {parquet_path} {BEELINK_HOST}:{beelink_path}/",
        shell=True,
    )
    if result.returncode != 0:
        print(f"rsync failed for {table_name}", file=sys.stderr)
        return False
    print(f"  Pushed to {BEELINK_HOST}:{beelink_path}/{parquet_path.name}")
    return True


def main():
    TEMP_DIR.mkdir(parents=True, exist_ok=True)

    ok = True
    ok &= write_and_push(fetch_estados_rows(), "estados")
    ok &= write_and_push(fetch_municipios_rows(), "municipios")

    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
