#!/usr/bin/env python3
"""
Fetch ANP (Agencia Nacional do Petroleo) weekly fuel resale price survey
("Levantamento de Precos de Combustiveis") -> Parquet -> beelink.

Source: www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/precos/
levantamento-de-precos-de-combustiveis-ultimas-semanas-pesquisadas -- the
HTML page itself is NOT WAF-blocked (plain GET, no special headers needed).
The individual per-week .xlsx download links ARE blocked with a plain
request (403, iso-8859-1 WAF error page) -- but adding a `Referer` header
pointing back at the same listing page makes the WAF let the download
through (confirmed via curl: identical request, 403 without Referer, 200
with it). This is a different failure mode than the old note ("403 on all
data pages") -- the page loads fine now, only the file download needed the
extra header.

Each weekly file is "revendas_lpc_<start>_<end>.xlsx": one row per gas
station per fuel product surveyed that week (CNPJ, razao social, endereco,
municipio, estado, bandeira, produto, unidade, preco de revenda, data da
coleta). Real header row starts after 9 rows of report preamble.

The listing page currently has weekly files going back to 2022 (~187 files
total). To keep this a quick, bounded pull, this script only fetches the
CURRENT CALENDAR YEAR's weekly files by default (pass a --year to override,
or --all-years for the full history, which is much slower: ~190 files at
~15-20s/file due to throttled downloads).

Usage:
    python3 scripts/scrap/anp_combustiveis.py [--year YYYY] [--all-years]
"""

import argparse
import re
import subprocess
import sys
import time
from datetime import date
from pathlib import Path

import requests

BEELINK_HOST = "beelink"
BEELINK_PATH = "~/rodado/br_anp_combustiveis/precos"

PAGE_URL = (
    "https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/"
    "precos/levantamento-de-precos-de-combustiveis-ultimas-semanas-pesquisadas"
)
UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"
TEMP_DIR = Path(
    "/private/tmp/claude-501/-Users-polux-Projetos-rodado/c780c9c0-b6b3-44b0-964e-08a3b2f2024c/scratchpad/anp"
)

LINK_RE = re.compile(r'href="(https://www\.gov\.br/anp/[^"]*/revendas_lpc_(\d{4})-[^"]*\.xlsx)"')


def find_revendas_links(session: requests.Session) -> list[tuple[str, str]]:
    resp = session.get(PAGE_URL, timeout=30)
    resp.raise_for_status()
    return [(m.group(1), m.group(2)) for m in LINK_RE.finditer(resp.text)]


def parse_revendas_xlsx(xlsx_bytes: bytes) -> list[dict]:
    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    columns = None
    rows_out = []
    for row in ws.iter_rows(values_only=True):
        if columns is None:
            if row and row[0] == "CNPJ":
                columns = [
                    "cnpj", "razao", "fantasia", "endereco", "numero", "complemento",
                    "bairro", "cep", "municipio", "estado", "bandeira", "produto",
                    "unidade_medida", "preco_revenda", "data_coleta",
                ]
            continue
        if row[0] is None:
            continue
        record = {}
        for col_name, value in zip(columns, row):
            if value is None:
                record[col_name] = None
            elif col_name == "preco_revenda":
                try:
                    record[col_name] = float(value)
                except (TypeError, ValueError):
                    record[col_name] = None
            else:
                # Normalize everything else (cnpj, numero, cep, data_coleta, ...)
                # to string -- source column types are inconsistent row-to-row
                # (e.g. CNPJ sometimes numeric, sometimes text with punctuation).
                record[col_name] = str(value)
        rows_out.append(record)
    wb.close()
    return rows_out


def main():
    import pyarrow as pa
    import pyarrow.parquet as pq

    parser = argparse.ArgumentParser()
    parser.add_argument("--year", default=None, help="Only fetch weeks from this year (default: current year)")
    parser.add_argument("--all-years", action="store_true", help="Fetch full available history (slow)")
    args = parser.parse_args()

    TEMP_DIR.mkdir(parents=True, exist_ok=True)

    session = requests.Session()
    session.headers.update({"User-Agent": UA, "Referer": PAGE_URL})

    print("Fetching ANP listing page to discover current weekly file links...")
    links = find_revendas_links(session)
    print(f"Found {len(links)} weekly 'revendas' files total")
    if not links:
        print("No download links found on the page -- aborting.")
        return 1

    if not args.all_years:
        target_year = args.year or str(date.today().year)
        links = [(url, yr) for url, yr in links if yr == target_year]
        print(f"Filtered to year {target_year}: {len(links)} files")

    all_rows = []
    for i, (url, yr) in enumerate(links):
        print(f"[{i + 1}/{len(links)}] Downloading {url} ...")
        resp = session.get(url, timeout=60)
        if resp.status_code != 200:
            print(f"  HTTP {resp.status_code} -- skipping")
            continue
        rows = parse_revendas_xlsx(resp.content)
        print(f"  {len(resp.content) / 1e6:.1f} MB, {len(rows)} rows")
        all_rows.extend(rows)
        time.sleep(0.5)

    print(f"\nTotal rows: {len(all_rows)}")
    if not all_rows:
        print("No rows fetched -- aborting, not pushing an empty file.")
        return 1

    table = pa.Table.from_pylist(all_rows)
    parquet_path = TEMP_DIR / "precos.parquet"
    pq.write_table(table, str(parquet_path), compression="zstd")
    print(f"Wrote {parquet_path} ({parquet_path.stat().st_size / 1e6:.1f} MB, {table.num_rows} rows)")

    subprocess.run(f"ssh {BEELINK_HOST} 'mkdir -p {BEELINK_PATH}'", shell=True, check=True)
    result = subprocess.run(
        f"rsync -av {parquet_path} {BEELINK_HOST}:{BEELINK_PATH}/",
        shell=True,
    )
    if result.returncode != 0:
        print("rsync failed", file=sys.stderr)
        return 1

    print(f"Pushed to {BEELINK_HOST}:{BEELINK_PATH}/{parquet_path.name}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
