#!/usr/bin/env python3
"""
Fetch CMED (Camara de Regulacao do Mercado de Medicamentos) price-ceiling
lists -> Parquet -> beelink.

Source: www.gov.br/anvisa/pt-br/assuntos/medicamentos/cmed/precos -- unlike
several other gov.br pages hit this session (ANVISA "consulta completa",
RENAME, PGFN, FNDE, ANP, ...), this specific page is NOT behind the gov.br
WAF: plain `requests` with a normal browser User-Agent gets a clean HTTP 200
both for the HTML page and the two xlsx download links it contains. The page
links to two lists, both updated periodically (looked ~monthly, but the
file's date-stamped filename changes more often than that in practice):

  - "xls_conformidade_site_<timestamp>.xlsx" -> PMC list (Preco Fabrica /
    Preco Maximo ao Consumidor -- what pharmacies can charge)
  - "xls_conformidade_gov_<timestamp>.xlsx"  -> PMVG list (Preco Maximo de
    Venda ao Governo -- what public-sector buyers can be charged; the more
    corruption-relevant of the two for detecting overpriced public
    purchases)

Both files embed a timestamp in the filename that changes on every
publish, so this script re-scrapes the HTML page on each run to find the
current links rather than hardcoding a URL. Each xlsx has ~30 rows of
legal/explanatory preamble before the real header row (auto-detected here
by scanning for the row whose first cell is "SUBSTÂNCIA"); the data table
itself is wide (~79 columns: substance, CNPJ, lab, GGREM code, registry,
product, presentation, PF/PMC broken out per ICMS-rate bracket, etc).
Column names are sanitized (accents/spaces/%/parens stripped) for
Parquet/DuckDB compatibility. Both lists share the same shape and are
concatenated with a `lista` tag column ("pmc" / "pmvg").

Usage:
    python3 scripts/scrap/anvisa_cmed.py
"""

import io
import re
import subprocess
import sys
import unicodedata
from pathlib import Path

import requests

BEELINK_HOST = "beelink"
BEELINK_PATH = "~/rodado/br_anvisa_cmed/precos"

PAGE_URL = "https://www.gov.br/anvisa/pt-br/assuntos/medicamentos/cmed/precos"
UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"
TEMP_DIR = Path(
    "/private/tmp/claude-501/-Users-polux-Projetos-rodado/c780c9c0-b6b3-44b0-964e-08a3b2f2024c/scratchpad/cmed"
)

LINK_RE = re.compile(
    r'href="(https://www\.gov\.br/anvisa/pt-br/assuntos/medicamentos/cmed/precos/arquivos/'
    r'xls_conformidade_(site|gov)_[^"]+\.xlsx/@@download/file)"'
)


def sanitize_column(name) -> str:
    if name is None:
        return "col_unnamed"
    s = str(name).strip()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = s.lower()
    s = s.replace("%", "pct")
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = s.strip("_")
    return s or "col_unnamed"


def find_download_links(session: requests.Session) -> dict:
    resp = session.get(PAGE_URL, timeout=30)
    resp.raise_for_status()
    links = {}
    for m in LINK_RE.finditer(resp.text):
        url, kind = m.group(1), m.group(2)
        # kind "site" == PMC (consumer-facing list), "gov" == PMVG (government-sale list)
        key = "pmc" if kind == "site" else "pmvg"
        links[key] = url
    return links


def parse_price_list(xlsx_bytes: bytes, lista: str) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header = None
    columns = None
    rows_out = []
    for row in ws.iter_rows(values_only=True):
        if header is None:
            if row and row[0] and str(row[0]).strip().upper() == "SUBSTÂNCIA":
                header = row
                columns = [sanitize_column(c) for c in row]
            continue
        if row[0] is None and row[1] is None:
            continue
        record = {"lista": lista}
        for col_name, value in zip(columns, row):
            record[col_name] = str(value) if value is not None else None
        rows_out.append(record)
    wb.close()
    return rows_out


def main():
    import pyarrow as pa
    import pyarrow.parquet as pq

    TEMP_DIR.mkdir(parents=True, exist_ok=True)

    session = requests.Session()
    session.headers.update({"User-Agent": UA})

    print("Fetching CMED precos page to discover current download links...")
    links = find_download_links(session)
    print(f"Found links: {links}")
    if not links:
        print("No download links found on the page -- aborting.")
        return 1

    all_rows = []
    for lista, url in links.items():
        print(f"Downloading {lista} list: {url}")
        resp = session.get(url, timeout=120)
        resp.raise_for_status()
        print(f"  {len(resp.content) / 1e6:.1f} MB")
        rows = parse_price_list(resp.content, lista)
        print(f"  parsed {len(rows)} rows")
        all_rows.extend(rows)

    print(f"\nTotal rows: {len(all_rows)}")
    if not all_rows:
        print("No rows fetched -- aborting, not pushing an empty file.")
        return 1

    table = pa.Table.from_pylist(all_rows)
    parquet_path = TEMP_DIR / "precos.parquet"
    pq.write_table(table, str(parquet_path), compression="zstd")
    print(f"Wrote {parquet_path} ({parquet_path.stat().st_size / 1e6:.1f} MB, {table.num_rows} rows)")

    subprocess.run(f"ssh {BEELINK_HOST} 'mkdir -p {BEELINK_PATH}'", shell=True, check=True)
    result = subprocess.run(
        f"rsync -av {parquet_path} {BEELINK_HOST}:{BEELINK_PATH}/",
        shell=True,
    )
    if result.returncode != 0:
        print("rsync failed", file=sys.stderr)
        return 1

    print(f"Pushed to {BEELINK_HOST}:{BEELINK_PATH}/{parquet_path.name}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
