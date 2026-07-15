#!/usr/bin/env python3
"""
Fetch SINAPI (Sistema Nacional de Pesquisa de Custos e Indices da Construcao
Civil) insumo prices -> Parquet -> beelink.

Source: Caixa Economica Federal's public downloads SharePoint site,
www.caixa.gov.br. The human-facing page (site/paginas/downloads.aspx) loads
its file list client-side via a SharePoint REST API
(_api/web/lists/getbytitle(...)/Items) -- discovered by reading
/Style Library/js/downloads.js, which is the JS the download page itself
uses. No auth needed, but every request (including the actual zip download)
requires a `security=true` cookie that Azion's edge sets on the *first*
request to any caixa.gov.br page and then expects back on every subsequent
request -- without it, both the HTML pages and the zip downloads 302-redirect
to themselves forever. So this script "warms up" the cookie with one GET
before doing anything else.

Since 2025 the monthly report has been published as ONE national zip per
month (category "SINAPI - Relatorios mensais - a partir de 2025" in the
Downloads list) containing 4 xlsx files, rather than 27 separate per-state
files as in earlier years (those live in 27 separate "ate 2024 - <UF>"
categories with a different, older layout -- out of scope here, see notes
below). The national zip's "SINAPI_Referencia_<year>_<month>.xlsx" workbook
has one sheet per encargo/desoneracao combination; this script uses sheet
"ISD" (Relatorio de Precos de Insumos - Encargos Sociais SEM Desoneracao),
the standard reference most public-works budgets cite. Each sheet is a wide
table with one column per UF (27 columns) that gets melted into long rows:
(mes_referencia, classificacao, codigo_insumo, descricao_insumo, unidade,
origem_preco, uf, preco_mediano).

Scope: only the "a partir de 2025" category is fetched (in practice this
currently covers 2024-10 through the latest published month -- Caixa
apparently backfilled a bit before 2025 into this bucket too). That's a
~21-month window, comfortably within the "last 24 months" target from the
task brief. Pre-2024 monthly reports use the old per-state-zip layout and
are NOT fetched by this script (would need a second, differently-shaped
fetch path per UF x month x year -- much larger and lower value than the
recent window).

Usage:
    python3 scripts/scrap/caixa_sinapi.py
"""

import io
import re
import subprocess
import sys
import time
import zipfile
from pathlib import Path

import requests

BEELINK_HOST = "beelink"
BEELINK_PATH = "~/rodado/br_caixa_sinapi/insumos"

BASE = "https://www.caixa.gov.br"
CATEGORIES_URL = (
    BASE + "/_api/web/lists/getbytitle('LT_T077_Downloads_Categorias')/Items"
    "?$select=ID,Title&$top=5000&$orderby=Title"
)
DOWNLOADS_URL_TMPL = (
    BASE + "/_api/web/lists/getbytitle('Downloads')/Items"
    "?$select=Title,Modified,File_x0020_Type,FileLeafRef,EncodedAbsUrl,"
    "Descricao,FileSizeDisplay,Categoria/ID"
    "&$expand=Categoria"
    "&$filter=Categoria/ID eq {cat_id} and FSObjType eq 0 and OData__ModerationStatus eq 0"
    "&$top=5000&$orderby=Modified desc"
)
UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"
TEMP_DIR = Path(
    "/private/tmp/claude-501/-Users-polux-Projetos-rodado/c780c9c0-b6b3-44b0-964e-08a3b2f2024c/scratchpad/sinapi"
)
UF_COLUMNS = [
    "AC", "AL", "AM", "AP", "BA", "CE", "DF", "ES", "GO", "MA", "MG", "MS",
    "MT", "PA", "PB", "PE", "PI", "PR", "RJ", "RN", "RO", "RR", "RS", "SC",
    "SE", "SP", "TO",
]


def make_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({"User-Agent": UA, "Accept": "application/json;odata=verbose"})
    # Warm up the Azion edge "security=true" cookie -- without it every
    # request (HTML or file) 302-redirects to itself forever.
    s.get(BASE + "/site/paginas/downloads.aspx", timeout=30)
    if "security" not in s.cookies.get_dict():
        print("  warning: security cookie not set after warmup, continuing anyway")
    return s


def find_current_category(session: requests.Session) -> tuple[int, str]:
    resp = session.get(CATEGORIES_URL, timeout=30)
    resp.raise_for_status()
    results = resp.json()["d"]["results"]
    candidates = [
        r for r in results
        if r["Title"].startswith("SINAPI - Relatórios mensais - a partir de")
    ]
    if not candidates:
        raise RuntimeError("no 'SINAPI - Relatórios mensais - a partir de ...' category found")
    # If more than one such bucket ever exists, prefer the most recent (highest ID).
    best = max(candidates, key=lambda r: r["ID"])
    return best["ID"], best["Title"]


def list_month_zips(session: requests.Session, cat_id: int) -> list[dict]:
    resp = session.get(DOWNLOADS_URL_TMPL.format(cat_id=cat_id), timeout=30)
    resp.raise_for_status()
    results = resp.json()["d"]["results"]
    by_month = {}
    for r in results:
        name = r["FileLeafRef"]
        m = re.search(r"SINAPI-(\d{4})-(\d{2})-formato-xlsx", name, re.IGNORECASE)
        if not m:
            continue
        month_key = f"{m.group(1)}-{m.group(2)}"
        existing = by_month.get(month_key)
        if existing is None or r["Modified"] > existing["Modified"]:
            by_month[month_key] = r
    return [by_month[k] for k in sorted(by_month.keys())]


def download_zip(session: requests.Session, url: str) -> bytes:
    for attempt in range(4):
        resp = session.get(url, timeout=90)
        if resp.status_code == 200 and resp.content[:2] == b"PK":
            return resp.content
        print(f"    attempt {attempt + 1}: status={resp.status_code} size={len(resp.content)}, retrying...")
        time.sleep(3 + attempt * 2)
    raise RuntimeError(f"failed to download a valid zip from {url}")


def parse_referencia_sheet(xlsx_bytes: bytes, mes_referencia: str) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    if "ISD" not in wb.sheetnames:
        print(f"    ISD sheet missing for {mes_referencia}, sheets={wb.sheetnames}")
        return []
    ws = wb["ISD"]
    rows_out = []
    header_idx = None
    header = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if header_idx is None:
            if row and row[1] and "Código" in str(row[1]):
                header_idx = i
                header = row
            continue
        if row[1] is None:
            continue
        classificacao, codigo, descricao, unidade, origem = row[0], row[1], row[2], row[3], row[4]
        for col_idx, uf in enumerate(UF_COLUMNS, start=5):
            if col_idx >= len(row):
                break
            preco = row[col_idx]
            if preco is None:
                continue
            rows_out.append({
                "mes_referencia": mes_referencia,
                "classificacao": classificacao,
                "codigo_insumo": codigo,
                "descricao_insumo": descricao,
                "unidade": unidade,
                "origem_preco": origem,
                "uf": uf,
                "preco_mediano": float(preco) if isinstance(preco, (int, float)) else None,
            })
    wb.close()
    return rows_out


def main():
    import pyarrow as pa
    import pyarrow.parquet as pq

    TEMP_DIR.mkdir(parents=True, exist_ok=True)

    session = make_session()

    cat_id, cat_title = find_current_category(session)
    print(f"Using category {cat_id}: {cat_title!r}")

    months = list_month_zips(session, cat_id)
    print(f"Found {len(months)} monthly zips: {[m['FileLeafRef'] for m in months]}")

    # Keep at most the most recent 24 months, per the task's "reasonable
    # recent window" guidance (this category currently has ~21, so normally
    # a no-op).
    months = months[-24:]

    all_rows = []
    for item in months:
        name = item["FileLeafRef"]
        url = item["EncodedAbsUrl"].replace("http://", "https://")
        m = re.search(r"SINAPI-(\d{4})-(\d{2})", name)
        mes_referencia = f"{m.group(1)}-{m.group(2)}"
        print(f"Fetching {name} ({mes_referencia}) ...")
        try:
            zip_bytes = download_zip(session, url)
        except Exception as e:
            print(f"  FAILED to download {name}: {e}")
            continue

        zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
        member = None
        for n in zf.namelist():
            if re.search(r"Refer.ncia", n):
                member = n
                break
        if member is None:
            print(f"  no 'Referência' xlsx member found in {name}, skipping")
            continue

        xlsx_bytes = zf.read(member)
        rows = parse_referencia_sheet(xlsx_bytes, mes_referencia)
        print(f"  parsed {len(rows)} rows from {member}")
        all_rows.extend(rows)

    print(f"\nTotal rows: {len(all_rows)}")
    if not all_rows:
        print("No rows fetched -- aborting, not pushing an empty file.")
        return 1

    table = pa.Table.from_pylist(all_rows)
    parquet_path = TEMP_DIR / "insumos.parquet"
    pq.write_table(table, str(parquet_path), compression="zstd")
    print(f"Wrote {parquet_path} ({parquet_path.stat().st_size / 1e6:.1f} MB, {table.num_rows} rows)")

    subprocess.run(f"ssh {BEELINK_HOST} 'mkdir -p {BEELINK_PATH}'", shell=True, check=True)
    result = subprocess.run(
        f"rsync -av {parquet_path} {BEELINK_HOST}:{BEELINK_PATH}/",
        shell=True,
    )
    if result.returncode != 0:
        print("rsync failed", file=sys.stderr)
        return 1

    print(f"Pushed to {BEELINK_HOST}:{BEELINK_PATH}/{parquet_path.name}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
